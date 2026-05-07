#!/usr/bin/env python3
from openpyxl import load_workbook

wb = load_workbook('update/books_classification.xlsx')
ws = wb.active

print('Excel Validation Summary:')
print(f'  Total books: {ws.max_row - 1}')
print(f'  Total columns: {ws.max_column}')

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print(f'\n  Columns: {headers}')

# Count classified vs unclassified
classified = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 6).value)
unclassified = ws.max_row - 1 - classified

print(f'\n  Classified: {classified}')
print(f'  Unclassified: {unclassified}')

# Count cross-refs
cross_refs = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 3).value)
print(f'  Cross-references found: {cross_refs}')

# Count sources
lib_json = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == 'library.json')
det_json = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == 'library-details.json')
print(f'\n  Source split:')
print(f'    library.json: {lib_json}')
print(f'    library-details.json: {det_json}')

print('\nFile location: /Users/jzuluaga/dev/jorgezuluaga/update/books_classification.xlsx')
