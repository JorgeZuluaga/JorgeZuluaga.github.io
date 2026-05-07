#!/usr/bin/env python3
"""
Export all books (library.json + library-details.json) to an Excel file
for manual Dewey classification.

Output: update/books_classification.xlsx
Columns: Source | bookId/ISBN | Title | Author | DCC Class columns | Reasoning | Confidence
"""

import json
import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
LIB_JSON = ROOT / "info" / "library.json"
DET_JSON = ROOT / "info" / "library-details.json"
OUT_XLSX = ROOT / "update" / "books_classification.xlsx"

DEWEY_AREA_NAMES = {
    0:   "Generalidades",
    100: "Filosofia y psicologia",
    200: "Religion",
    300: "Ciencias sociales",
    400: "Lenguas",
    500: "Ciencias naturales y matematicas",
    600: "Tecnologia",
    700: "Artes y recreacion",
    800: "Literatura",
    900: "Historia y geografia",
}


def area_name(code: int) -> str:
    base = math.floor(code / 100) * 100
    return DEWEY_AREA_NAMES.get(base, str(base))


def extract_dcc_classes(dcc_classes, dcc_codes) -> list[str]:
    """Return normalized class labels like ['Ciencias sociales (300)', 'Lenguas (400)']."""
    parts = []
    seen = set()

    if isinstance(dcc_classes, dict):
        for code_str in dcc_classes:
            try:
                code = int(code_str)
            except ValueError:
                continue
            base = math.floor(code / 100) * 100
            if base not in seen:
                seen.add(base)
                parts.append(f"{area_name(base)} ({base})")

    if isinstance(dcc_classes, list):
        for item in dcc_classes:
            try:
                code = int(item)
            except (TypeError, ValueError):
                continue
            base = math.floor(code / 100) * 100
            if base not in seen:
                seen.add(base)
                parts.append(f"{area_name(base)} ({base})")

    if isinstance(dcc_codes, dict):
        code_iter = dcc_codes.keys()
    elif isinstance(dcc_codes, list):
        code_iter = dcc_codes
    else:
        code_iter = []

    for code_str in code_iter:
        try:
            code = int(code_str)
        except (TypeError, ValueError):
            continue
        base = math.floor(code / 100) * 100
        if base not in seen:
            seen.add(base)
            parts.append(f"{area_name(base)} ({base})")

    return parts


def extract_notes(dcc_notes) -> tuple[str, str]:
    """Return (reasoning, confidence) from dcc_notes when available."""
    if not isinstance(dcc_notes, dict):
        return "", ""

    reasoning = dcc_notes.get("reasoning", "")
    confidence = dcc_notes.get("confidence", "")
    confidence_text = "" if confidence in (None, "") else str(confidence)
    return str(reasoning or ""), confidence_text


def display_bookid(source: str, book_id: str, isbn: str) -> str:
    if source == "library-details.json":
        return isbn or book_id
    return book_id


def max_class_count(rows: list[dict]) -> int:
    return max((len(r["dcc_classes_list"]) for r in rows), default=0)


def base_headers_and_widths(class_cols: int) -> tuple[list[str], list[int]]:
    headers = ["Fuente", "bookId/ISBN", "Título", "Autor"]
    widths = [18, 24, 50, 30]

    for idx in range(1, class_cols + 1):
        headers.append(f"Clase DCC {idx} (actual)")
        widths.append(34)

    headers.extend(["Razonamiento", "Confianza", "Clases DCC (manual)"])
    widths.extend([64, 14, 45])
    return headers, widths


def row_values(row: dict, class_cols: int) -> list[str]:
    values = [
        row["source"],
        display_bookid(row["source"], row["bookId"], row.get("isbn", "")),
        row["title"],
        row["author"],
    ]

    classes = row["dcc_classes_list"]
    for idx in range(class_cols):
        values.append(classes[idx] if idx < len(classes) else "")

    values.extend([
        row.get("reasoning", ""),
        row.get("confidence", ""),
        "",  # manual column — empty for user to fill
    ])
    return values


def load_library_books() -> list[dict]:
    data = json.loads(LIB_JSON.read_text(encoding="utf-8"))
    rows = []
    for b in data.get("books", []):
        if not b or not b.get("title"):
            continue
        reasoning, confidence = extract_notes(b.get("dcc_notes"))
        rows.append({
            "source": "library.json",
            "bookId": str(b.get("bookId", "")),
            "isbn": str(b.get("ISBN", b.get("isbn", "")) or ""),
            "title": b.get("title", ""),
            "author": b.get("author", ""),
            "dcc_classes_list": extract_dcc_classes(b.get("dcc_classes"), b.get("dcc_codes")),
            "reasoning": reasoning,
            "confidence": confidence,
        })
    return rows


def load_details_books(library_ids: set) -> list[dict]:
    data = json.loads(DET_JSON.read_text(encoding="utf-8"))
    rows = []
    for b in data.get("books", []):
        if not b or not b.get("Title"):
            continue
        book_id = str(b.get("bookId", ""))
        isbn = str(b.get("ISBN", b.get("isbn", "")) or "")
        # Skip books already in library.json
        if book_id and book_id in library_ids:
            continue
        reasoning, confidence = extract_notes(b.get("dcc_notes"))
        rows.append({
            "source": "library-details.json",
            "bookId": book_id,
            "isbn": isbn,
            "title": b.get("Title", ""),
            "author": b.get("Author", ""),
            "dcc_classes_list": extract_dcc_classes(b.get("dcc_classes"), b.get("dcc_codes")),
            "reasoning": reasoning,
            "confidence": confidence,
        })
    return rows


def build_rows() -> list[dict]:
    lib_rows = load_library_books()
    lib_ids = {str(r["bookId"]) for r in lib_rows if r["bookId"]}
    det_rows = load_details_books(lib_ids)
    all_rows = lib_rows + det_rows
    # Sort: unclassified first, then alphabetical by title
    all_rows.sort(key=lambda r: (bool(r["dcc_classes_list"]), r["title"].lower()))
    return all_rows


def write_excel(rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libros"

    class_cols = max_class_count(rows)
    headers, col_widths = base_headers_and_widths(class_cols)

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # Alternating row fills
    fill_even = PatternFill("solid", fgColor="EFF6FF")
    fill_odd = PatternFill("solid", fgColor="FFFFFF")
    fill_classified = PatternFill("solid", fgColor="F0FDF4")

    wrap = Alignment(vertical="top", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        fill = fill_classified if row["dcc_classes_list"] else (fill_even if row_idx % 2 == 0 else fill_odd)
        values = row_values(row, class_cols)
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = wrap

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Saved {len(rows)} books → {OUT_XLSX}")
    classified = sum(1 for r in rows if r["dcc_classes_list"])
    print(f"  Classified: {classified} | Unclassified: {len(rows) - classified}")


if __name__ == "__main__":
    rows = build_rows()
    write_excel(rows)
